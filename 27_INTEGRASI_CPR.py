import firebase_admin
from firebase_admin import credentials, db
import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from tkinter import messagebox
from tkinter.ttk import Progressbar
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from ttkbootstrap.scrolled import ScrolledText
import threading
import time
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os 
from openpyxl.chart import ScatterChart, Reference, Series 
from openpyxl.chart.axis import ChartLines 

# Import skfuzzy
import numpy as np
import skfuzzy as fuzz
from skfuzzy import control as ctrl

# === Firebase Init ===
try:
    cred = credentials.Certificate("data-comunication-test.json")
    firebase_admin.initialize_app(cred, {
        'databaseURL': 'https://data-comunication-test-default-rtdb.asia-southeast1.firebasedatabase.app/'
    })
except Exception as e:
    root = ttk.Window()
    root.withdraw()
    messagebox.showerror("Error Kredensial", f"File 'data-comunication-test.json' tidak ditemukan. Pastikan file tersebut ada di direktori yang sama dengan skrip.\n\nError: {e}")
    exit()


logs_ref = db.reference("/CPR_LOGS")
status_ref = db.reference("/CPR/status")
summary_ref = db.reference("/CPR")

gui_started = False
status_text = "🕒 WAITING"
progress_value = 0
session_start_wib = None 
session_end_wib = None   

# === FUZZY LOGIC SYSTEM DEFINITION ===
# Input 1: Kedalaman CPR
# Crisp Value: 0-9 cm
kedalaman = ctrl.Antecedent(np.arange(0, 10.1, 0.1), 'kedalaman') # 0-9 cm

# Input 2: Ritme CPR
# Crisp Value: 50-150 cpm
ritme = ctrl.Antecedent(np.arange(0, 151, 1), 'ritme') # 0-150 cpm

# Output: Feedback Realtime (nilai)
# Crisp Value: 30 - 100
feedback = ctrl.Consequent(np.arange(0, 101, 1), 'feedback') 

# Fungsi Keanggotaan (Membership Functions) untuk Input
# Kedalaman CPR
# Terlalu Dangkal (TD): 0-5 cm
kedalaman['terlalu_dangkal'] = fuzz.trimf(kedalaman.universe, [0, 0, 5]) 
# Cukup (C): 4-7 cm
kedalaman['cukup'] = fuzz.trimf(kedalaman.universe, [4, 5.5, 7])      
# Terlalu Dalam (TTD): 6-9 cm
kedalaman['terlalu_dalam'] = fuzz.trimf(kedalaman.universe, [6, 9, 9]) 

# Ritme CPR
# Terlalu Lambat (TL): 0-100 cpm
ritme['terlalu_lambat'] = fuzz.trimf(ritme.universe, [0, 0, 100])     
# Ideal (I): 95-125 cpm
ritme['ideal'] = fuzz.trimf(ritme.universe, [95, 110, 125])       
# Terlalu Cepat (TC): 120-150 cpm
ritme['terlalu_cepat'] = fuzz.trimf(ritme.universe, [120, 150, 150]) 

# Fungsi Keanggotaan (Membership Functions) untuk Output
# Feedback Realtime (nilai)
# Perbaiki Kedalaman & Ritme (PKR): 30-60
feedback['perbaiki_kedalaman_ritme'] = fuzz.trimf(feedback.universe, [30, 30, 60]) 
# Perbaiki Kedalaman (PK): 50-70
feedback['perbaiki_kedalaman'] = fuzz.trimf(feedback.universe, [50, 60, 70])     
# Perbaiki Ritme (PR): 60-80
feedback['perbaiki_ritme'] = fuzz.trimf(feedback.universe, [60, 70, 80])         
# Bagus & Lanjutkan (BL): 80-100
feedback['bagus_lanjutkan'] = fuzz.trimf(feedback.universe, [80, 100, 100])      

# Aturan Fuzzy (Rule Base)
# Variabel Linguistik | Ritme TL | Ritme I | Ritme TC
# --------------------|----------|---------|-----------
# Kedalaman TD        | PKR      | PK      | PKR
# Kedalaman C         | PR       | BL      | PR
# Kedalaman TTD       | PKR      | PK      | PKR

rule1 = ctrl.Rule(kedalaman['terlalu_dangkal'] & ritme['terlalu_lambat'], feedback['perbaiki_kedalaman_ritme'])
rule2 = ctrl.Rule(kedalaman['terlalu_dangkal'] & ritme['ideal'], feedback['perbaiki_kedalaman'])
rule3 = ctrl.Rule(kedalaman['terlalu_dangkal'] & ritme['terlalu_cepat'], feedback['perbaiki_kedalaman_ritme'])

rule4 = ctrl.Rule(kedalaman['cukup'] & ritme['terlalu_lambat'], feedback['perbaiki_ritme'])
rule5 = ctrl.Rule(kedalaman['cukup'] & ritme['ideal'], feedback['bagus_lanjutkan'])
rule6 = ctrl.Rule(kedalaman['cukup'] & ritme['terlalu_cepat'], feedback['perbaiki_ritme'])

rule7 = ctrl.Rule(kedalaman['terlalu_dalam'] & ritme['terlalu_lambat'], feedback['perbaiki_kedalaman_ritme'])
rule8 = ctrl.Rule(kedalaman['terlalu_dalam'] & ritme['ideal'], feedback['perbaiki_kedalaman'])
rule9 = ctrl.Rule(kedalaman['terlalu_dalam'] & ritme['terlalu_cepat'], feedback['perbaiki_kedalaman_ritme'])

# Sistem Kontrol Fuzzy
feedback_ctrl = ctrl.ControlSystem([rule1, rule2, rule3, rule4, rule5, rule6, rule7, rule8, rule9])
fuzzy_simulator = ctrl.ControlSystemSimulation(feedback_ctrl)

def calculate_fuzzy_score(avg_depth, cpm_last_value):
    """
    Menghitung skor fuzzy berdasarkan rata-rata kedalaman dan nilai CPM terakhir.
    """
    try:
        # Pastikan input berada dalam universe yang didefinisikan
        # Clip nilai agar tidak melebihi batas universe
        avg_depth_clipped = np.clip(avg_depth, kedalaman.universe.min(), kedalaman.universe.max())
        cpm_last_value_clipped = np.clip(cpm_last_value, ritme.universe.min(), ritme.universe.max())

        fuzzy_simulator.input['kedalaman'] = avg_depth_clipped
        fuzzy_simulator.input['ritme'] = cpm_last_value_clipped
        fuzzy_simulator.compute()
        return round(fuzzy_simulator.output['feedback'], 2)
    except Exception as e:
        print(f"Error during fuzzy computation: {e}")
        # Jika terjadi error (misal input di luar range universe dan tidak di-clip), kembalikan nilai default
        return 0 

# === Ambil data dari Firebase CPR_LOGS ===
def ambil_data():
    snapshot = logs_ref.get()
    if not snapshot:
        return pd.DataFrame()
    df = pd.DataFrame.from_dict(snapshot, orient='index')
    df.index = pd.to_numeric(df.index)
    df.sort_index(inplace=True)
    return df

def simpan_ke_excel():
    global session_start_wib, session_end_wib

    df = ambil_data()
    if df.empty:
        messagebox.showwarning("Data Kosong", "⚠️ Tidak ada data untuk disimpan.")
        return

    # --- Data Filtering and Processing ---
    # 1. Hapus data dengan cpm = 0
    df_filtered_cpm = df[df['cpm'] != 0].copy()

    if df_filtered_cpm.empty:
        messagebox.showwarning("Data Kosong", "⚠️ Tidak ada data valid setelah filter (CPM > 0).")
        return

    # 2. Sortir data double CPM dengan nilai kedalaman terbesar
    df_sorted_cpm_depth = df_filtered_cpm.sort_values(by=['cpm', 'kedalaman_cm'], ascending=[True, False])
    df_processed = df_sorted_cpm_depth.drop_duplicates(subset=['cpm'], keep='first')

    # --- Calculate Averages and Last CPM from Processed Data ---
    # Rata-rata kedalaman yang dihitung hanya kedalaman diatas 3cm
    df_depth_above_3_for_avg_depth = df_processed[df_processed['kedalaman_cm'] > 4].copy() # Filter > 3cm
    avg_k_for_fuzzy = round(df_depth_above_3_for_avg_depth["kedalaman_cm"].mean(), 2) if not df_depth_above_3_for_avg_depth.empty else 2.34

    avg_g_sorted = round(df_processed["gaya_N"].mean(), 2)
    
    # CPM terakhir dari data yang sudah diproses dan difilter cpm != 0
    cpm_last_value_for_fuzzy = int(df_processed["cpm"].iloc[-1]) if not df_processed.empty else 0
    
    # Hitung skor fuzzy
    skor_fuzzy = calculate_fuzzy_score(avg_k_for_fuzzy, cpm_last_value_for_fuzzy)
   
    # Persiapan data untuk disimpan
    df_final_excel = df_processed.reset_index().rename(columns={"index": "timestamp_ms"})
    
    if session_start_wib:
        df_final_excel["waktu"] = df_final_excel["timestamp_ms"].apply(
            lambda ms: (session_start_wib + timedelta(milliseconds=ms)).strftime("%H:%M:%S.%f")[:-3]
        )
    else:
        df_final_excel["waktu"] = df_final_excel["timestamp_ms"].apply(
            lambda x: f"{x//60000:02}:{(x%60000)//1000:02}.{x%1000:03}"
        )

    waktu_simpan = datetime.now().strftime("%Y%m%d_%H%M%S")
    nama_user = user_var.get().strip().replace(" ", "_")
    if not nama_user:
        nama_user = "User" 
    nama_file = f"CPR_{nama_user}_{waktu_simpan}.xlsx"

    try:
        # Menyimpan data mentah
        with pd.ExcelWriter(nama_file, engine='openpyxl') as writer:
            # Sheet Data CPR
            df_final_excel[['waktu', 'cpm', 'gaya_N', 'kedalaman_cm']].to_excel(writer, index=False, sheet_name='Data CPR', startrow=2) 

            workbook = writer.book
            worksheet_data = workbook['Data CPR']

            thin_border = Border(left=Side(style='thin'), 
                                 right=Side(style='thin'), 
                                 top=Side(style='thin'), 
                                 bottom=Side(style='thin'))

            worksheet_data['A1'] = f"Nama User: {user_var.get()}"
            worksheet_data.merge_cells('A1:D1') 
            worksheet_data['A1'].font = Font(bold=True)
            worksheet_data['A1'].alignment = Alignment(horizontal='left', vertical='center') 

            duration_str = "N/A"
            if session_start_wib and session_end_wib:
                duration = session_end_wib - session_start_wib
                total_seconds = int(duration.total_seconds())
                days = total_seconds // (24 * 3600)
                total_seconds %= (24 * 3600)
                hours = total_seconds // 3600
                total_seconds %= 3600
                minutes = total_seconds // 60
                seconds = total_seconds % 60
                
                duration_parts = []
                if days > 0:
                    duration_parts.append(f"{days} hari")
                if hours > 0:
                    duration_parts.append(f"{hours} jam")
                if minutes > 0:
                    duration_parts.append(f"{minutes} menit")
                if seconds > 0 or not duration_parts: 
                    duration_parts.append(f"{seconds} detik")
                duration_str = " ".join(duration_parts)
                
            worksheet_data['A2'] = f"Waktu Latihan: {duration_str.strip()}"
            worksheet_data.merge_cells('A2:D2')
            worksheet_data['A2'].font = Font(bold=True)
            worksheet_data['A2'].alignment = Alignment(horizontal='left', vertical='center')

            header_row_num = 3 
            for col_idx in range(1, 5): 
                cell = worksheet_data.cell(row=header_row_num, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            last_data_row = len(df_final_excel) + header_row_num 
            last_row_for_footer = last_data_row + 1 

            worksheet_data[f'A{last_row_for_footer}'] = "Rata-Rata"
            worksheet_data[f'A{last_row_for_footer}'].font = Font(bold=True)
            worksheet_data[f'A{last_row_for_footer}'].border = thin_border
            
            cpm_cell = worksheet_data[f'B{last_row_for_footer}']
            cpm_cell.value = cpm_last_value_for_fuzzy 
            cpm_cell.font = Font(bold=True)
            cpm_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
            cpm_cell.border = thin_border
            
            gaya_cell = worksheet_data[f'C{last_row_for_footer}']
            gaya_cell.value = avg_g_sorted 
            gaya_cell.font = Font(bold=True)
            gaya_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
            gaya_cell.border = thin_border
            
            # Kedalaman Rata-Rata (hanya > 3cm)
            kedalaman_cell = worksheet_data[f'D{last_row_for_footer}']
            kedalaman_cell.value = avg_k_for_fuzzy 
            kedalaman_cell.font = Font(bold=True)
            kedalaman_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") 
            kedalaman_cell.border = thin_border


            skor_cpr_row = last_row_for_footer + 1
            worksheet_data[f'A{skor_cpr_row}'] = "SKOR CPR (Fuzzy)"
            worksheet_data[f'A{skor_cpr_row}'].font = Font(bold=True)
            worksheet_data[f'A{skor_cpr_row}'].border = thin_border

            skor_fuzzy_cell = worksheet_data[f'B{skor_cpr_row}']
            skor_fuzzy_cell.value = skor_fuzzy
            skor_fuzzy_cell.font = Font(bold=True)
            skor_fuzzy_cell.fill = PatternFill(start_color="ADFF2F", end_color="ADFF2F", fill_type="solid") 
            skor_fuzzy_cell.border = thin_border

            worksheet_data[f'C{skor_cpr_row}'].border = thin_border 
            worksheet_data[f'D{skor_cpr_row}'].border = thin_border


            summary_data = {
                'Parameter': ['Nama User', 'Waktu Simpan', 'Waktu Latihan', 'Rata-Rata Kedalaman (cm)', 'Rata-Rata Gaya (N)', 'CPM Terakhir', 'SKOR CPR (Fuzzy)'], # Update label
                'Nilai': [user_var.get(), datetime.now().strftime("%Y-%m-%d %H:%M:%S"), duration_str.strip(), avg_k_for_fuzzy, avg_g_sorted, cpm_last_value_for_fuzzy, skor_fuzzy]
            }
            df_summary = pd.DataFrame(summary_data)
            df_summary.to_excel(writer, index=False, sheet_name='Ringkasan')

            worksheet_summary = workbook['Ringkasan']
            for col_idx in range(1, 3): 
                cell = worksheet_summary.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center') 
                
            for row_idx in range(2, len(df_summary) + 2): 
                for col_idx in range(1, 3):
                    cell = worksheet_summary.cell(row=row_idx, column=col_idx)
                    cell.border = thin_border
                    if col_idx == 2: 
                        cell.alignment = Alignment(horizontal='left', vertical='center')

            # --- Membuat Sheet untuk Data Grafik Excel (Tabel Data) ---
            # Data untuk chart akan diambil dari sini
            worksheet_chart_data_table = workbook.create_sheet('Data Grafik Tabel')
            worksheet_chart_data_table['A1'] = "Waktu (ms)"
            worksheet_chart_data_table['B1'] = "Kedalaman (cm)"
            worksheet_chart_data_table['C1'] = "Ritme (CPM)"
            
            # Apply bold and border to headers in Data Grafik Tabel
            for col_idx in range(1, 4):
                cell = worksheet_chart_data_table.cell(row=1, column=col_idx)
                cell.font = Font(bold=True)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Write data for charts
            for r_idx, row in df_final_excel.iterrows():
                row_num = r_idx + 2 # Data starts from row 2
                worksheet_chart_data_table.cell(row=row_num, column=1, value=row['timestamp_ms']).border = thin_border
                worksheet_chart_data_table.cell(row=row_num, column=2, value=row['kedalaman_cm']).border = thin_border
                worksheet_chart_data_table.cell(row=row_num, column=3, value=row['cpm']).border = thin_border
            
            # --- Membuat Sheet untuk Grafik Chart ---
            worksheet_chart = workbook.create_sheet('Grafik CPR')

            # Chart Kedalaman
            chart_depth = ScatterChart()
            chart_depth.title = "Kedalaman Kompresi (cm)"
            chart_depth.x_axis.title = "Waktu (ms)"
            chart_depth.y_axis.title = "Kedalaman (cm)"
            
            # Set chart size (width, height)
            # To extend width to column 'W', increase width value
            chart_depth.width = 30.0  # Adjusted width to extend to column W
            chart_depth.height = 10.0 # Keep original height

            # Tambahkan gridlines ke sumbu Y
            chart_depth.y_axis.majorGridlines = ChartLines()

            # Tentukan rentang data untuk chart Kedalaman
            # Pastikan max_row sesuai dengan jumlah data + baris header
            max_row_chart_data = len(df_final_excel) + 1 
            x_values_depth = Reference(worksheet_chart_data_table, min_col=1, min_row=2, max_row=max_row_chart_data)
            y_values_depth = Reference(worksheet_chart_data_table, min_col=2, min_row=2, max_row=max_row_chart_data)
            series_depth = Series(y_values_depth, x_values_depth, title="Kedalaman")
            chart_depth.series.append(series_depth)
            
            # Atur posisi chart di sheet
            worksheet_chart.add_chart(chart_depth, "A1") 

            # Chart Ritme
            chart_ritme = ScatterChart()
            chart_ritme.title = "Ritme Kompresi (CPM)"
            chart_ritme.x_axis.title = "Waktu (ms)"
            chart_ritme.y_axis.title = "Ritme (CPM)"

            # Set chart size (width, height)
            # To extend width to column 'W', increase width value
            chart_ritme.width = 30.0  # Adjusted width to extend to column W
            chart_ritme.height = 10.0 # Keep original height

            # Tambahkan gridlines ke sumbu Y
            chart_ritme.y_axis.majorGridlines = ChartLines()

            # Tentukan rentang data untuk chart Ritme
            x_values_ritme = Reference(worksheet_chart_data_table, min_col=1, min_row=2, max_row=max_row_chart_data)
            y_values_ritme = Reference(worksheet_chart_data_table, min_col=3, min_row=2, max_row=max_row_chart_data)
            series_ritme = Series(y_values_ritme, x_values_ritme, title="Ritme")
            chart_ritme.series.append(series_ritme)

            # Atur posisi chart di sheet (di bawah chart kedalaman)
            worksheet_chart.add_chart(chart_ritme, "A30") 

            # Auto-adjust column width for all sheets - FIX UNTUK MERGED CELL ERROR
            for sheet_name in workbook.sheetnames:
                worksheet = workbook[sheet_name]
                for col_idx in range(1, worksheet.max_column + 1):
                    column_letter = get_column_letter(col_idx)
                    max_length = 0
                    for row_idx in range(1, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        # Lewati jika sel adalah bagian dari sel yang digabungkan atau tidak memiliki nilai
                        if cell.value is not None and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                            try:
                                cell_value_str = str(cell.value)
                                if len(cell_value_str) > max_length:
                                    max_length = len(cell_value_str)
                            except Exception as e:
                                # Tangani kasus di mana konversi ke string gagal
                                print(f"Warning: Could not get length of cell value at {column_letter}{row_idx}: {e}")
                                pass
                        # Juga pertimbangkan header jika ada
                        elif row_idx == 1 and cell.value is not None:
                             try:
                                 cell_value_str = str(cell.value)
                                 if len(cell_value_str) > max_length:
                                     max_length = len(cell_value_str)
                             except:
                                 pass # abaikan jika gagal
                                 
                    adjusted_width = (max_length + 2)
                    if adjusted_width < 10: adjusted_width = 10 
                    worksheet.column_dimensions[column_letter].width = adjusted_width
            
        log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] ✅ File Excel disimpan: {nama_file}\n")
        log_box.see("end")
        
        # --- Tambahan: Otomatis membuka file Excel ---
        if os.path.exists(nama_file):
            try:
                os.startfile(nama_file) 
                log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] 🚀 Membuka file Excel: {nama_file}\n")
                log_box.see("end")
            except Exception as open_e:
                messagebox.showwarning("Gagal Membuka File", f"Gagal membuka file Excel secara otomatis. Anda dapat membukanya secara manual di: {os.path.abspath(nama_file)}\n\nError: {open_e}")
                log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] ❌ Gagal membuka file Excel secara otomatis.\n")
                log_box.see("end")
        else:
            messagebox.showwarning("File Tidak Ditemukan", f"File Excel tidak ditemukan setelah disimpan: {os.path.abspath(nama_file)}")

    except Exception as e:
        messagebox.showerror("Gagal Simpan", f"Terjadi kesalahan saat menyimpan: {e}")
        print(f"Error detail: {e}") 


def synchronize_time():
    global session_start_wib 
    session_start_wib = datetime.now()
    log_box.insert("end", f"[{session_start_wib.strftime('%H:%M:%S.%f')[:-3]}] ⌚ Waktu disinkronkan. Siap memulai sesi.\n")
    log_box.see("end")
    btn_sync_time.config(state="disabled") 
    btn_start.config(state="normal") 
    btn_reset.config(state="normal") # Aktifkan tombol reset setelah sinkronisasi

def update_logging():
    global gui_started, seen_timestamps, session_end_wib
    seen_timestamps = set()
    start_time = None
    
    while True:
        try:
            status = status_ref.get()

            if gui_started:
                if status == "Logging dimulai...":
                    if start_time is None:
                        start_time = time.time()
                    progress_var.set(min(time.time() - start_time, 60))
                    status_label.config(text="🟠 LOGGING")

                    snapshot = logs_ref.get()
                    if snapshot:
                        for ts_str, data in snapshot.items():
                            if ts_str not in seen_timestamps:
                                seen_timestamps.add(ts_str)
                                ts = int(ts_str) 
                                
                                waktu_str = f"{ts//60000:02}:{(ts%60000)//1000:02}.{ts%1000:03}"
                                
                                gaya = round(data.get("gaya_N", 0), 2)
                                kedalaman = round(data.get("kedalaman_cm", 0), 2)
                                cpm = int(data.get("cpm", 0))
                                log_box.insert("end", f"[{waktu_str}] 📌 Kedalaman: {kedalaman:.2f} cm | Gaya: {gaya:.2f} N | CPM: {cpm}\n")
                                log_box.see("end")

                        df_sorted = pd.DataFrame.from_dict(snapshot, orient='index')
                        df_sorted.index = pd.to_numeric(df_sorted.index, errors='coerce')
                        df_sorted = df_sorted.dropna(subset=["cpm", "kedalaman_cm", "gaya_N"]) 
                        df_sorted = df_sorted.astype({"cpm": int, "kedalaman_cm": float, "gaya_N": float}).sort_index()

                        df_graph = df_sorted[df_sorted['cpm'] != 0].copy()

                        if not df_graph.empty:
                            ax1.clear()
                            ax2.clear()
                            ax1.plot(df_graph.index, df_graph["kedalaman_cm"], color='blue', label="Kedalaman (cm)")
                            ax1.set_title("Kedalaman Kompresi")
                            ax1.set_ylabel("cm")
                            ax1.grid(True)
                            ax2.plot(df_graph.index, df_graph["cpm"], color='orange', label="Ritme (CPM)")
                            ax2.set_title("Ritme Kompresi (CPM)")
                            ax2.set_ylabel("CPM")
                            ax2.set_xlabel("Waktu (ms)")
                            canvas.draw()
                        else:
                            ax1.clear()
                            ax2.clear()
                            ax1.set_title("Kedalaman Kompresi")
                            ax2.set_title("Ritme Kompresi (CPM)")
                            ax1.set_ylabel("cm")
                            ax2.set_ylabel("CPM")
                            ax2.set_xlabel("Waktu (ms)")
                            canvas.draw() 
                                                                    
                elif status == "Logging selesai...":
                    session_end_wib = datetime.now() 
                    df = ambil_data()
                    df_filtered_summary = df[df['cpm'] != 0].copy()
                    
                    if not df_filtered_summary.empty:
                        df_sorted_cpm_depth_summary = df_filtered_summary.sort_values(by=['cpm', 'kedalaman_cm'], ascending=[True, False])
                        df_processed_summary = df_sorted_cpm_depth_summary.drop_duplicates(subset=['cpm'], keep='first')
                        
                        # Rata-rata kedalaman yang dihitung hanya kedalaman diatas 3cm
                        df_depth_above_3_summary = df_processed_summary[df_processed_summary['kedalaman_cm'] > 4].copy()
                        avg_k_summary = round(df_depth_above_3_summary["kedalaman_cm"].mean(), 2) if not df_depth_above_3_summary.empty else 2.34

                        avg_g_summary = round(df_processed_summary["gaya_N"].mean(), 2)
                        cpm_f_summary = int(df_processed_summary["cpm"].iloc[-1]) if not df_processed_summary.empty else 0
                        
                        skor_fuzzy_summary = calculate_fuzzy_score(avg_k_summary, cpm_f_summary)
                        
                        log_box.insert("end", f"\n[{datetime.now().strftime('%H:%M:%S')}] ✅ Logging selesai.\n")
                        log_box.insert("end", f"📊 Rata-rata: Kedalaman = {avg_k_summary} cm | Gaya = {avg_g_summary} N | CPM Terakhir = {cpm_f_summary}\n") 
                        log_box.insert("end", f"⭐ SKOR CPR (Fuzzy) = {skor_fuzzy_summary}\n\n")
                        log_box.see("end")
                        status_label.config(text="🟢 SELESAI")
                        
                        btn_save.config(state="normal") 
                    else:
                        log_box.insert("end", f"\n[{datetime.now().strftime('%H:%M:%S')}] ✅ Logging selesai, namun tidak ada data valid (CPM > 0) untuk dianalisis.\n")
                        log_box.see("end")
                        status_label.config(text="🟢 SELESAI")
                        btn_save.config(state="disabled") 
                        
                    gui_started = False
                    start_time = None
                    progress_var.set(0)
                    
                    status_ref.set("Menunggu Sesi Baru") 
                    btn_sync_time.config(state="normal") 
                    btn_reset.config(state="normal") # Aktifkan tombol reset setelah logging selesai
            
            elif not gui_started and status_label.cget("text") != "🟢 SELESAI":
                    status_label.config(text="🕒 WAITING")

        except Exception as e:
            log_box.insert("end", f"⚠️ Terjadi error pada background thread: {e}\n")
            log_box.see("end")
            print(f"Error in update_logging: {e}") 
        
        time.sleep(1)


def mulai_logging_gui():
    global gui_started, seen_timestamps, session_start_wib, session_end_wib
    if not user_var.get().strip():
        messagebox.showwarning("Nama Kosong", "⚠️ Silakan isi nama user terlebih dahulu.")
        return
    
    if session_start_wib is None:
        messagebox.showwarning("Waktu Belum Disinkronkan", "⚠️ Silakan klik 'Sinkronisasi Waktu' terlebih dahulu.")
        return

    try:
        # Hapus data dari Firebase untuk sesi baru yang bersih
        logs_ref.delete()
        summary_ref.delete() 
        status_ref.set("Menunggu Perintah")
        seen_timestamps = set()
        session_end_wib = None 
        
        # Bersihkan GUI
        log_box.delete('1.0', END)
        ax1.clear()
        ax2.clear()
        canvas.draw()
        log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] 🗑️ Data lama dihapus dari Firebase.\n")
        log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] ✅ Sesi baru dimulai oleh {user_var.get()}.\n")
        log_box.insert("end", f"📡 Menunggu perintah 'Logging dimulai...' dari perangkat IoT...\n")
        log_box.see("end")
        
        status_label.config(text="🕒 WAITING")
        btn_save.config(state="disabled")
        btn_sync_time.config(state="disabled") 
        btn_reset.config(state="disabled") # Nonaktifkan tombol reset saat logging berlangsung
        gui_started = True
    except Exception as e:
        messagebox.showerror("Firebase Error", f"Gagal menghapus data lama: {e}")

def reset_session():
    global gui_started, session_start_wib, session_end_wib, seen_timestamps
    
    confirm = messagebox.askyesno("Konfirmasi Reset", "Anda yakin ingin mereset sesi? Ini akan menghapus semua data di Firebase dan membersihkan GUI.")
    if not confirm:
        return

    try:
        # Reset Firebase
        logs_ref.delete()
        summary_ref.delete()
        status_ref.set("Menunggu Sesi Baru") # Kembali ke status awal
        
        # Reset GUI
        user_var.set("") # Kosongkan username
        log_box.delete('1.0', END) # Bersihkan logbox
        ax1.clear() # Bersihkan grafik
        ax2.clear()
        canvas.draw()
        
        # Reset variabel status
        gui_started = False
        session_start_wib = None
        session_end_wib = None
        seen_timestamps = set()
        progress_var.set(0)
        status_label.config(text="🕒 WAITING")
        
        # Atur ulang status tombol
        btn_sync_time.config(state="normal")
        btn_start.config(state="disabled")
        btn_save.config(state="disabled")
        btn_reset.config(state="disabled") # Nonaktifkan lagi sampai waktu disinkronkan
        
        log_box.insert("end", f"[{datetime.now().strftime('%H:%M:%S')}] 🔄 Sesi telah direset. Siap untuk sesi baru.\n")
        log_box.see("end")

    except Exception as e:
        messagebox.showerror("Reset Error", f"Terjadi kesalahan saat mereset sesi: {e}")


# === GUI Layout ===
app = ttk.Window("CPR Logger Realtime [Python + Firebase + IoT Cloud]", themename="minty")
app.state('zoomed')
app.protocol("WM_DELETE_WINDOW", app.destroy)

main_frame = ttk.Frame(app, padding=10)
main_frame.pack(fill='both', expand=True)

top_bar = ttk.Frame(main_frame)
top_bar.pack(fill='x', pady=5)

ttk.Label(top_bar, text="Status:", font=("Segoe UI", 11)).pack(side='left', padx=(0, 5))
status_label = ttk.Label(top_bar, text=status_text, font=("Segoe UI", 11, "bold"), bootstyle="info", width=15)
status_label.pack(side='left', padx=5)

ttk.Label(top_bar, text="Nama User:", font=("Segoe UI", 11)).pack(side='left', padx=(20, 5))
user_var = ttk.StringVar()
entry_user = ttk.Entry(top_bar, textvariable=user_var, font=("Segoe UI", 11), width=30)
entry_user.pack(side='left', padx=5)
entry_user.insert(0, "Fabian") # Nilai default

btn_sync_time = ttk.Button(top_bar, text="⌚ SINKRONISASI WAKTU", bootstyle="warning", command=synchronize_time)
btn_sync_time.pack(side='left', padx=10)

btn_start = ttk.Button(top_bar, text="▶ MULAI SESI BARU", bootstyle="success", command=mulai_logging_gui, state="disabled") 
btn_start.pack(side='left', padx=10)

btn_reset = ttk.Button(top_bar, text="🔄 RESET SESI", bootstyle="danger", command=reset_session, state="disabled") # Tombol Reset
btn_reset.pack(side='left', padx=10)

btn_save = ttk.Button(top_bar, text="💾 SIMPAN KE EXCEL", bootstyle="primary", state='disabled', command=simpan_ke_excel)
btn_save.pack(side='right', padx=5)

progress_var = ttk.DoubleVar()
progress_bar = Progressbar(main_frame, maximum=60, variable=progress_var, bootstyle="info-striped")
progress_bar.pack(fill='x', pady=10)

content_frame = ttk.Frame(main_frame)
content_frame.pack(fill='both', expand=True)
content_frame.columnconfigure(1, weight=1)
content_frame.rowconfigure(0, weight=1)

log_box = ScrolledText(content_frame, width=60, font=("Consolas", 10), autohide=True, relief="solid", borderwidth=1)
log_box.grid(row=0, column=0, sticky='nsew', padx=(0, 5))

fig = plt.figure(figsize=(8, 5))
ax1 = fig.add_subplot(211)
ax2 = fig.add_subplot(212)
fig.tight_layout(pad=3.0)
canvas = FigureCanvasTkAgg(fig, master=content_frame)
canvas_widget = canvas.get_tk_widget()
canvas_widget.grid(row=0, column=1, sticky='nsew')

threading.Thread(target=update_logging, daemon=True).start()
log_box.insert("end", "🩺 GUI Siap. Masukkan nama, klik 'Sinkronisasi Waktu', lalu 'MULAI SESI BARU'.\n")
app.mainloop()
